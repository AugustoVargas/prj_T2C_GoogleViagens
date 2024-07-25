from botcity.web import WebBot, Browser,By
from botcity.core import DesktopBot
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CMaestro import T2CMaestro, LogLevel, ErrorType
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CExceptions import BusinessRuleException
from prj_T2C_GoogleViagens.classes_t2c.sqlite.T2CSqliteQueue import T2CSqliteQueue
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime
import re, os
import pandas as pd
import openpyxl
from openpyxl import load_workbook


class FuncoesGoogleViagens:
    def __init__(self, arg_dictConfig:dict, arg_clssMaestro:T2CMaestro, arg_botWebbot:WebBot=None, arg_botDesktopbot:DesktopBot=None,arg_clssSqliteQueue:T2CSqliteQueue=None):
        '''
        Classe com funções que serão utilizadas para navegação e realização de operações dentro do portal Google Viagens   

        Parâmetros:
        - arg_dictConfig (dict): dicionário de configuração.
        - arg_clssMaestro (T2CMaestro): instância de T2CMaestro.
        - arg_botWebbot (WebBot): instância de WebBot (opcional, default=None).
        - arg_botDesktopbot (DesktopBot): instância de DesktopBot (opcional, default=None).
        - arg_clssSqliteQueue (T2CSqliteQueue): instância de T2CSqliteQueue (opcional, default=None).
        
        '''
        if(arg_botWebbot is None and arg_botDesktopbot is None): raise Exception("Não foi possível inicializar a classe, forneça pelo menos um bot")
        else:
            self.var_botWebbot = arg_botWebbot
            self.var_botDesktopbot = arg_botDesktopbot
            self.var_dictConfig = arg_dictConfig
            self.var_clssMaestro = arg_clssMaestro
            self.var_clssSqliteQueue = arg_clssSqliteQueue

    def abrirGoogleViagens(self, arg_strURLGoogleViagens):
        """
        Função que realiza o acesso ao site GoogleViagens
        
        Parâmetros:
            - arg_strURLDadosMundiais: Url de acesso ao site GoogleViagens        
        Retorna:
        """
        try:                    
            #realiza o acesso ao site GoogleViagens
            self.var_clssMaestro.write_log("Abrindo o portal GoogleViagens")
           
            #abre o site do GoogleViagens
            self.var_botWebbot.browse(arg_strURLGoogleViagens)
            self.var_botWebbot.maximize_window()

            var_intCont = 0
            var_xpathPaginaInicialGV = self.var_botWebbot.find_element('.//a[@aria-label="Google"]', By.XPATH)            

            #aguardando carregamento do site
            while not var_xpathPaginaInicialGV and var_intCont<10:
                    self.var_clssMaestro.write_log("carregando o site GoogleViagens, aguarde.....")
                    self.var_botWebbot.wait(10000)    
                    var_intCont = var_intCont + 1

                    #verifica se o contador de tentativas é igual a 10
                    if var_intCont >= 10:
                        raise Exception("falha ao carregar o site GoogleViagens")
                    
            self.var_clssMaestro.write_log("GoogleViagens, carregado com sucesso!")       

          

        except Exception as exception:              
            raise


    def selecionarIdaGoogleViagens(self):
        """
        Função que realiza a seleção do modo da viagem no site GoogleViagens
        
        Parâmetros:
                   
        Retorna:
        """
        try:                    
            
            self.var_clssMaestro.write_log("Selecionando o modo da viagem no GoogleViagens")

            #seleciona o modo de viagem           
            var_xpathModoviagem= self.var_botWebbot.find_element('.//div[@class="RLVa8 GeHXyb"]', By.XPATH)    
            var_xpathModoviagem.click()        

            self.var_botWebbot.wait(2000)  
            #seleciona a opção Ida e volta
            var_xpathIdaeVolta = self.var_botWebbot.find_element('.//li[@class="MCs1Pd UbEQCe VfPpkd-OkbHre VfPpkd-OkbHre-SfQLQb-M1Soyc-bN97Pc VfPpkd-aJasdd-RWgCYc-wQNmvb ib1Udf VfPpkd-rymPhb-ibnC6b VfPpkd-rymPhb-ibnC6b-OWXEXe-gk6SMd VfPpkd-rymPhb-ibnC6b-OWXEXe-SfQLQb-M1Soyc-Bz112c VfPpkd-rymPhb-ibnC6b-OWXEXe-SfQLQb-Woal0c-RWgCYc"]',By.XPATH)
            var_xpathModoviagem.click() 
                                
            self.var_clssMaestro.write_log("Modo viagem selecionado com sucesso!")            

        except Exception as exception:              
            raise


    def selecionarSaida(self):
        """
        Função que realiza a seleção do ponto inicial da viagem no site GoogleViagens
        
        Parâmetros:
                   
        Retorna:
        """
        try:                    
            #inicializa o nome da cidade de origem
            var_strCidade = self.var_dictConfig["strCidadeOrigem"]
            self.var_clssMaestro.write_log("Selecionando o ponto inicial da viagem no GoogleViagens")

            #seleciona o ponto de partida         
            var_wemPontoPartida= self.var_botWebbot.find_element('.//input[@aria-label="De onde?"]', By.XPATH)    
            var_wemPontoPartida.click()  
            self.var_botWebbot.wait(2000)  

            #preenche o ponto de partida
            var_wemPartida = self.var_botWebbot.find_element('.//input[@aria-label="Outros pontos de partida?"]',By.XPATH)
            var_wemPartida.clear()
            var_wemPartida.send_keys(var_strCidade)
            self.var_botWebbot.wait(3000)  

            #seleciona a cidade na lista suspensa
            var_wemCidade = self.var_botWebbot.find_element(f'.//li[@aria-label="{var_strCidade}"]',By.XPATH)
            var_wemCidade.click()                               
            self.var_clssMaestro.write_log("Ponto de origem da viagem selecionado com sucesso!")            

        except Exception as exception:              
            raise

    def selecionarDestino(self, arg_strPaisDestino):
        """
        Função que realiza a seleção do Destino da viagem no site GoogleViagens
        
        Parâmetros:
                   
        Retorna:
        """
        try:                    
            #inicializa o nome da cidade de origem
            
            self.var_clssMaestro.write_log("Selecionando o Destino da viagem no GoogleViagens")
            self.var_botWebbot.wait(2000)  
            #seleciona o ponto de Destino         
            var_wemPontoDestino= self.var_botWebbot.find_element('.//input[@aria-label="Para onde?"and @aria-autocomplete="inline"]', By.XPATH)    
            var_wemPontoDestino.click()  
            self.var_botWebbot.wait(2000)  
        
            #preenche o ponto de Destino
            var_wemDestino = self.var_botWebbot.find_element('.//input[@aria-label="Para onde?" and @aria-expanded="true"]',By.XPATH)
            var_wemDestino.clear()
            var_wemDestino.send_keys(arg_strPaisDestino)
            self.var_botWebbot.wait(3000)  

            #seleciona o país na lista suspensa
            var_wemPais = self.var_botWebbot.find_element(f'.//li[@aria-label="{arg_strPaisDestino}"]',By.XPATH)
            var_wemPais.click()                               
            self.var_clssMaestro.write_log("Ponto de Destino da viagem selecionado com sucesso!")            

        except Exception as exception:              
            raise
    
    def selecionaPeriodo(self):

        """
        Função que realiza a seleção do período da viagem no site GoogleViagens
        
        Parâmetros:
                   
        Retorna:
        """
        try:                    
            # inicializa o período da viagem
            var_data_embarque = datetime.datetime.now()
            var_data_embarque_str = var_data_embarque.strftime("%d/%m/%Y")

            # define a data de retorno para 3 dias após a data de embarque
            var_data_retorno = var_data_embarque + datetime.timedelta(days=3)
            var_data_retorno_str = var_data_retorno.strftime("%d/%m/%Y")

            
            
            self.var_clssMaestro.write_log("Selecionando o peíodo da viagem no GoogleViagens")
            self.var_botWebbot.wait(2000)  

            #seleciona o período             
            if self.var_botWebbot.find( "clickViagem", matching=0.97, waiting_time=10000):               
                self.var_botWebbot.click()
            self.var_botWebbot.wait(5000)  

            self.var_clssMaestro.write_log("Selecionando Data específica")
            #seleciona a opção data específica
            if not self.var_botWebbot.find( "datasEspecificas", matching=0.97, waiting_time=10000):
                raise ("datasEspecificas")
            self.var_botWebbot.click()            
             
            self.var_botWebbot.wait(2000)
        
            self.var_clssMaestro.write_log("Preenchendo a data de embarque")
            #preenche a data de embarque
            var_wemEmbarque = self.var_botWebbot.find_element('.//input[@type="text" and @placeholder="Partida" and @aria-describedby="i28"]',By.XPATH)
            var_wemEmbarque.click() 
            var_wemEmbarque.clear()
            var_wemEmbarque.send_keys(var_data_embarque_str)
            self.var_botWebbot.wait(3000)  

            #preenche a data de retorno
            self.var_clssMaestro.write_log("Preenchendo a data de retorno")
            var_wemVolta = self.var_botWebbot.find_element('.//input[@type="text" and @placeholder="Volta" and @aria-describedby="i28"]',By.XPATH)
            var_wemVolta.click()   
            var_wemVolta.clear()
            var_wemVolta.send_keys(var_data_retorno_str)                          
            self.var_clssMaestro.write_log("Ponto de Destino da viagem selecionado com sucesso!")  
            var_wemclick=  self.var_botWebbot.find_element('.//div[@jsaction="JIbuQc:FhSeVc(McfNlf),xzUtuf(S9gUrf); click:npT2md;"]',By.XPATH)
            var_wemclick.click()

            
            self.var_botWebbot.wait(3000)

            self.var_clssMaestro.write_log("Concluindo o ajuste do período da viagem")
            #clicar no concluído
            var_wemBotaoConcluido =self.var_botWebbot.find_element('.//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ nCP5yc AjY5Oe DuMIQc LQeN7 z18xM Q74FEc"]', By.XPATH)
            var_wemBotaoConcluido.click()



        except Exception as exception:              
            raise


    def extraiInfosValores(self, arg_strPais):

        """
        Função que realiza a a extração dos valores das passagens no site GoogleViagens
        
        Parâmetros:
        -arg_strPais: recebe o nome do país para ser registrado na tabela de saída   
        Retorna:
        """
        try:
            self.var_clssMaestro.write_log("Extração dos valores das passagens viagem no GoogleViagens iniciada....")

            var_strCaminhoArquivoPassagens = self.var_dictConfig["strCaminhoArquivoPassagens"]
            var_strNomeArquivo = os.path.join(var_strCaminhoArquivoPassagens,"Passagens.xlsx")
            var_wbkResultado = load_workbook(var_strNomeArquivo)
            var_wbksheet = var_wbkResultado['Todos']
            self.var_botWebbot.wait(5000)
            var_intQtdCidades = self.var_botWebbot.execute_javascript(code='return document.getElementsByClassName("MJg7fb QB2Jof").length')
            
            for index in range(var_intQtdCidades):
                # Captura o nome da cidade
                var_strNomeCidade = self.var_botWebbot.execute_javascript(code=f'return document.getElementsByClassName("W6bZuc YMlIz")[{index}].textContent')
                self.var_botWebbot.wait(1000)
                # Captura o preço da passagem
                var_strPrecoPassagem = self.var_botWebbot.execute_javascript(code=f'return document.getElementsByClassName("MJg7fb QB2Jof")[{index}].innerText')
                
                print(var_strNomeCidade + var_strPrecoPassagem)

                # Determina a próxima linha vazia no arquivo Excel
                next_row = var_wbksheet.max_row + 1

                # Escreve os valores nas células correspondentes
                var_wbksheet.cell(row=next_row, column=1).value = arg_strPais
                var_wbksheet.cell(row=next_row, column=2).value = var_strNomeCidade
                var_wbksheet.cell(row=next_row, column=3).value = var_strPrecoPassagem

            # Salva o arquivo Excel após completar o loop
            var_wbkResultado.save(var_strNomeArquivo)
            self.var_clssMaestro.write_log("Informações registradas com Sucesso!")

        
        except Exception as exception:              
            raise


    def sortPassagensBaratas(self):

        """
        Função que realiza a seleção dos valores mais baratos das passagens extraídas do site GoogleViagens
        
        Parâmetros:
           
        Retorna:
        """
        try:
            self.var_clssMaestro.write_log("Filtragem do valores das passagens mais baratas iniciada....")

            var_strCaminhoArquivoPassagens = self.var_dictConfig["strCaminhoArquivoPassagens"]
            var_strNomeArquivo = os.path.join(var_strCaminhoArquivoPassagens,"Passagens.xlsx")

            # Carrega o arquivo Excel para um DataFrame do Pandas
            # Carrega o arquivo Excel para um DataFrame do Pandas
            df = pd.read_excel(var_strNomeArquivo, sheet_name='Todos')
            df['Preço'] = df['Preço'].str.replace('.', '').str.replace(',', '.').str.replace("R$","").astype(float)

            df_menoresValores = df.nsmallest(10,'Preço',keep='all')
            df.to_excel(var_strNomeArquivo, sheet_name='Todos')
            df_menoresValores.to_excel(var_strNomeArquivo, sheet_name='Baratos')
            
            self.var_clssMaestro.write_log("Filtragem do valores das passagens mais baratas concluída....")

        except Exception as exception:              
            raise